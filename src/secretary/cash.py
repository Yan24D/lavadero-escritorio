"""
Módulo de Cierre de Caja Diario para Secretario
Gestiona reportes diarios, estadísticas y exportación de datos
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import csv
from datetime import datetime
from database.db_config import db
from secretary.base_module import BaseModule


class CashModule(BaseModule):
    """Módulo para cierre de caja y reportes"""
    
    def __init__(self, user_data):
        super().__init__(user_data)
        self.records_tree = None
        self.current_date = datetime.now().strftime('%Y-%m-%d')
        self.summary_labels = {}
        self.lavador_stats_frame = None
        self.servicios_stats_frame = None
    
    def render(self, parent):
        """Renderizar módulo de cierre de caja"""
        self.parent_frame = parent
        self.clear_parent(parent)
        
        # Contenedor con scroll
        canvas = tk.Canvas(parent, bg='#f8fafc')
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f8fafc')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Botones de acción
        self.create_action_buttons(scrollable_frame)
        
        # Cards de resumen
        self.create_summary_cards(scrollable_frame)
        
        # Tabla de registros
        self.create_records_table(scrollable_frame)
        
        # Estadísticas
        self.create_analytics_section(scrollable_frame)
        
        self.load_data()
    
    def create_action_buttons(self, parent):
        """Crear botones de acción superiores"""
        actions_frame = tk.Frame(parent, bg='#f8fafc')
        actions_frame.pack(fill='x', pady=(0, 20))
        
        export_btn = self.create_button(
            actions_frame, "📊 Exportar CSV",
            self.export_csv, bg_color='#059669'
        )
        export_btn.pack(side='left', padx=(0, 10))
        
        report_btn = self.create_button(
            actions_frame, "🖨️ Imprimir Reporte",
            self.print_report, bg_color='#dc2626'
        )
        report_btn.pack(side='left', padx=(0, 10))
        
        refresh_btn = self.create_button(
            actions_frame, "🔄 Actualizar",
            self.refresh, bg_color='#0891b2'
        )
        refresh_btn.pack(side='right')
    
    def create_summary_cards(self, parent):
        """Crear cards de resumen financiero"""
        cards_frame = tk.Frame(parent, bg='#f8fafc')
        cards_frame.pack(fill='x', pady=(0, 20))
        
        card_data = [
            ("INGRESOS DEL DÍA", 'ingresos', '#3b82f6'),
            ("TOTAL COMISIONES", 'comisiones', '#6b7280'),
            ("BALANCE NETO", 'balance', '#059669'),
            ("SERVICIOS REALIZADOS", 'servicios', '#2563eb')
        ]
        
        for title, key, color in card_data:
            card_frame = tk.Frame(cards_frame, bg=color, relief='solid', borderwidth=1)
            card_frame.pack(side='left', fill='x', expand=True, padx=5)
            
            card_content = tk.Frame(card_frame, bg=color)
            card_content.pack(fill='both', expand=True, padx=20, pady=15)
            
            tk.Label(
                card_content, text=title,
                font=('Segoe UI', 10, 'bold'),
                fg='white', bg=color
            ).pack()
            
            value_label = tk.Label(
                card_content, text="$0" if 'S' not in title else "0",
                font=('Segoe UI', 18, 'bold'),
                fg='white', bg=color
            )
            value_label.pack(pady=(5, 0))
            
            self.summary_labels[key] = value_label
    
    def create_records_table(self, parent):
        """Crear tabla de registros del día"""
        table_container = tk.Frame(parent, bg='white', relief='solid', borderwidth=1)
        table_container.pack(fill='both', expand=True, pady=(0, 20))
        
        table_header = tk.Frame(table_container, bg='#f8fafc')
        table_header.pack(fill='x', padx=1, pady=1)
        
        header_content = tk.Frame(table_header, bg='#f8fafc')
        header_content.pack(fill='x', padx=15, pady=10)
        
        today_str = datetime.now().strftime('%A, %d de %B de %Y')
        table_title = tk.Label(
            header_content,
            text=f"📋 Registros del Día - {today_str}",
            font=('Segoe UI', 14, 'bold'),
            fg='#1e293b',
            bg='#f8fafc'
        )
        table_title.pack(side='left')
        
        columns = ('ID', 'Hora', 'Vehículo', 'Placa', 'Servicio',
                   'Costo', 'Comisión', 'Lavador', 'Estado')
        
        self.records_tree = ttk.Treeview(
            table_container, columns=columns, show='headings', height=12
        )

        # Menú contextual
        self.context_menu = tk.Menu(self.records_tree, tearoff=0)
        self.context_menu.add_command(label="✏️ Editar", command=self.edit_record)
        self.context_menu.add_command(label="🗑️ Eliminar", command=self.delete_record)
        
        # Bind del menú contextual
        self.records_tree.bind("<Button-3>", self.show_context_menu)  # Clic derecho
        self.records_tree.bind("<Double-1>", lambda e: self.edit_record())  # Doble clic
        
        for col in columns:
            self.records_tree.heading(col, text=col)
        
        widths = [50, 80, 100, 100, 120, 100, 100, 120, 100]
        for col, width in zip(columns, widths):
            self.records_tree.column(col, width=width, anchor='center')
        
        table_scrollbar = ttk.Scrollbar(
            table_container, orient='vertical', command=self.records_tree.yview
        )
        self.records_tree.configure(yscrollcommand=table_scrollbar.set)
        
        self.records_tree.pack(
            side='left', fill='both', expand=True, padx=(15, 0), pady=(0, 15)
        )
        table_scrollbar.pack(side='right', fill='y', pady=(0, 15), padx=(0, 15))
    
    def create_analytics_section(self, parent):
        """Crear sección de análisis"""
        analytics_frame = tk.Frame(parent, bg='#f8fafc')
        analytics_frame.pack(fill='x')
        
        # Rendimiento por lavador
        lavador_frame = tk.Frame(analytics_frame, bg='white', relief='solid', borderwidth=1)
        lavador_frame.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        lavador_content = tk.Frame(lavador_frame, bg='white')
        lavador_content.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(
            lavador_content,
            text="👥 Rendimiento por Lavador",
            font=('Segoe UI', 14, 'bold'),
            fg='#1e293b',
            bg='white'
        ).pack(anchor='w', pady=(0, 15))
        
        self.lavador_stats_frame = tk.Frame(lavador_content, bg='white')
        self.lavador_stats_frame.pack(fill='both', expand=True)
        
        # Servicios más solicitados
        servicios_frame = tk.Frame(analytics_frame, bg='white', relief='solid', borderwidth=1)
        servicios_frame.pack(side='right', fill='both', expand=True, padx=(10, 0))
        
        servicios_content = tk.Frame(servicios_frame, bg='white')
        servicios_content.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(
            servicios_content,
            text="⭐ Servicios Más Solicitados",
            font=('Segoe UI', 14, 'bold'),
            fg='#1e293b',
            bg='white'
        ).pack(anchor='w', pady=(0, 15))
        
        self.servicios_stats_frame = tk.Frame(servicios_content, bg='white')
        self.servicios_stats_frame.pack(fill='both', expand=True)
    
    def load_data(self):
        """Cargar datos del día"""
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            
            # Obtener resumen
            summary_query = """
                SELECT 
                    COUNT(*) as total_servicios,
                    COALESCE(SUM(costo), 0) as total_ingresos,
                    COALESCE(SUM(comision_calculada), 0) as total_comisiones,
                    COALESCE(SUM(ganancia_neta), 0) as balance_neto
                FROM vista_registros_completos
                WHERE fecha = %s
            """
            
            summary = db.execute_query(summary_query, (today,))
            
            if summary:
                data = summary[0]
                self.summary_labels['ingresos'].config(
                    text=f"${data['total_ingresos']:,.0f}"
                )
                self.summary_labels['comisiones'].config(
                    text=f"${data['total_comisiones']:,.0f}"
                )
                self.summary_labels['balance'].config(
                    text=f"${data['balance_neto']:,.0f}"
                )
                self.summary_labels['servicios'].config(
                    text=str(data['total_servicios'])
                )
            
            self.load_records(today)
            self.load_lavador_stats(today)
            self.load_servicios_stats(today)
            
        except Exception as e:
            print(f"Error cargando datos de caja: {e}")
            messagebox.showerror("Error", "Error al cargar datos del cierre de caja")
    
    def load_records(self, date):
        """Cargar registros del día"""
        try:
            for item in self.records_tree.get_children():
                self.records_tree.delete(item)
            
            query = """
                SELECT id, hora, vehiculo_nombre, placa, servicio_nombre,
                    costo, comision_calculada, lavador, pago
                FROM vista_registros_completos
                WHERE fecha = %s
                ORDER BY hora ASC
            """
            
            records = db.execute_query(query, (date,))
            
            if records:
                for record in records:
                    hora = self.format_time(record['hora'])
                    costo = self.format_currency(record['costo'])
                    comision = self.format_currency(record['comision_calculada'])
                    
                    self.records_tree.insert('', 'end', values=(
                        record['id'], hora, record['vehiculo_nombre'],
                        record['placa'], record['servicio_nombre'],
                        costo, comision, record['lavador'], record['pago']
                    ))
            else:
                self.records_tree.insert('', 'end', values=(
                    '', '', '', '', 'No hay registros para hoy',
                    '', '', '', ''
                ))
                
        except Exception as e:
            print(f"Error cargando registros: {e}")
    
    def load_lavador_stats(self, date):
        """Cargar estadísticas de lavadores"""
        try:
            for widget in self.lavador_stats_frame.winfo_children():
                widget.destroy()
            
            query = """
                SELECT lavador, COUNT(*) as servicios,
                    SUM(comision_calculada) as total_comision
                FROM vista_registros_completos
                WHERE fecha = %s
                GROUP BY lavador
                ORDER BY total_comision DESC
                LIMIT 5
            """
            
            stats = db.execute_query(query, (date,))
            
            if stats:
                for i, stat in enumerate(stats, 1):
                    row_frame = tk.Frame(self.lavador_stats_frame, bg='white')
                    row_frame.pack(fill='x', pady=2)
                    
                    pos_label = tk.Label(
                        row_frame, text=f"{i}.",
                        font=('Segoe UI', 10, 'bold'),
                        bg='white', fg='#374151', width=2
                    )
                    pos_label.pack(side='left')
                    
                    name_label = tk.Label(
                        row_frame, text=stat['lavador'],
                        font=('Segoe UI', 10),
                        bg='white', fg='#1e293b'
                    )
                    name_label.pack(side='left', padx=(5, 0))
                    
                    stats_label = tk.Label(
                        row_frame,
                        text=f"{stat['servicios']} servicios - ${stat['total_comision']:,.0f}",
                        font=('Segoe UI', 9),
                        bg='white', fg='#6b7280'
                    )
                    stats_label.pack(side='right')
            else:
                no_data = tk.Label(
                    self.lavador_stats_frame,
                    text="No hay datos de lavadores",
                    font=('Segoe UI', 10),
                    fg='#6b7280', bg='white'
                )
                no_data.pack(expand=True)
                
        except Exception as e:
            print(f"Error cargando estadísticas de lavadores: {e}")
    
    def load_servicios_stats(self, date):
        """Cargar estadísticas de servicios"""
        try:
            for widget in self.servicios_stats_frame.winfo_children():
                widget.destroy()
            
            query = """
                SELECT servicio_nombre, COUNT(*) as cantidad,
                    SUM(costo) as total_ingresos
                FROM vista_registros_completos
                WHERE fecha = %s
                GROUP BY servicio_nombre
                ORDER BY cantidad DESC
                LIMIT 5
            """
            
            stats = db.execute_query(query, (date,))
            
            if stats:
                for i, stat in enumerate(stats, 1):
                    row_frame = tk.Frame(self.servicios_stats_frame, bg='white')
                    row_frame.pack(fill='x', pady=2)
                    
                    pos_label = tk.Label(
                        row_frame, text=f"{i}.",
                        font=('Segoe UI', 10, 'bold'),
                        bg='white', fg='#374151', width=2
                    )
                    pos_label.pack(side='left')
                    
                    name_label = tk.Label(
                        row_frame, text=stat['servicio_nombre'],
                        font=('Segoe UI', 10),
                        bg='white', fg='#1e293b'
                    )
                    name_label.pack(side='left', padx=(5, 0))
                    
                    stats_label = tk.Label(
                        row_frame,
                        text=f"{stat['cantidad']} veces - ${stat['total_ingresos']:,.0f}",
                        font=('Segoe UI', 9),
                        bg='white', fg='#6b7280'
                    )
                    stats_label.pack(side='right')
            else:
                no_data = tk.Label(
                    self.servicios_stats_frame,
                    text="No hay datos de servicios",
                    font=('Segoe UI', 10),
                    fg='#6b7280', bg='white'
                )
                no_data.pack(expand=True)
                
        except Exception as e:
            print(f"Error cargando estadísticas de servicios: {e}")
    
    def export_csv(self):
        """Exportar datos a Excel (XLSX)"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Error", 
                "Módulo openpyxl no instalado.\n\n"
                "Ejecute: pip install openpyxl")
            return
        
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            
            query = """
                SELECT fecha, hora, vehiculo_nombre, placa, servicio_nombre,
                    costo, porcentaje, comision_calculada, ganancia_neta,
                    lavador, pago, usuario_nombre
                FROM vista_registros_completos
                WHERE fecha = %s
                ORDER BY hora ASC
            """
            
            data = db.execute_query(query, (today,))
            
            if not data:
                messagebox.showinfo("Información", "No hay datos para exportar")
                return
            
            filename = filedialog.asksaveasfilename(
                title="Guardar Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"cierre_caja_{today}.xlsx"
            )
            
            if filename:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Cierre de Caja"
                
                # Encabezados
                headers = ['Fecha', 'Hora', 'Vehículo', 'Placa', 'Servicio',
                        'Costo', 'Porcentaje', 'Comisión', 'Ganancia Neta',
                        'Lavador', 'Estado', 'Usuario']
                
                ws.append(headers)
                
                # Estilo de encabezados
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Datos
                for row in data:
                    ws.append([
                        row['fecha'], str(row['hora'])[:5] if row['hora'] else '',
                        row['vehiculo_nombre'], row['placa'], row['servicio_nombre'],
                        row['costo'], row['porcentaje'], row['comision_calculada'],
                        row['ganancia_neta'], row['lavador'], row['pago'],
                        row['usuario_nombre']
                    ])
                
                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column_letter].width = max_length + 2
                
                wb.save(filename)
                messagebox.showinfo("Éxito", f"Datos exportados a:\n{filename}")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar: {str(e)}")
    
    def print_report(self):
        """Generar reporte imprimible"""
        try:
            today = datetime.now().strftime('%Y-%m-%d')
            today_formatted = datetime.now().strftime('%A, %d de %B de %Y')
            
            summary_query = """
                SELECT 
                    COUNT(*) as total_servicios,
                    COALESCE(SUM(costo), 0) as total_ingresos,
                    COALESCE(SUM(comision_calculada), 0) as total_comisiones,
                    COALESCE(SUM(ganancia_neta), 0) as balance_neto
                FROM vista_registros_completos
                WHERE fecha = %s
            """
            
            summary = db.execute_query(summary_query, (today,))
            
            records_query = """
                SELECT hora, vehiculo_nombre, placa, servicio_nombre,
                    costo, lavador, pago
                FROM vista_registros_completos
                WHERE fecha = %s
                ORDER BY hora ASC
            """
            
            records = db.execute_query(records_query, (today,))
            
            preview_window = tk.Toplevel(self.parent_frame)
            preview_window.title("Vista Previa - Reporte Diario")
            preview_window.geometry("800x600")
            preview_window.configure(bg='white')
            
            text_frame = tk.Frame(preview_window)
            text_frame.pack(fill='both', expand=True, padx=20, pady=20)
            
            text_area = tk.Text(text_frame, font=('Courier New', 10), wrap='none')
            scrollbar_v = ttk.Scrollbar(text_frame, orient='vertical',
                command=text_area.yview)
            
            text_area.configure(yscrollcommand=scrollbar_v.set)
            
            report_content = f"""
{'='*80}
                        CLEAN CAR - REPORTE DIARIO
{'='*80}

Fecha: {today_formatted}
Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
Usuario: {self.user_data['nombre']}

{'='*80}
                            RESUMEN FINANCIERO
{'='*80}
"""
            
            if summary:
                data = summary[0]
                report_content += f"""
Total de Servicios:     {data['total_servicios']:>10}
Ingresos del Día:      ${data['total_ingresos']:>10,.0f}
Total Comisiones:      ${data['total_comisiones']:>10,.0f}
Balance Neto:          ${data['balance_neto']:>10,.0f}
"""
            
            report_content += f"""
{'='*80}
                        DETALLE DE SERVICIOS
{'='*80}

{'Hora':<8} {'Vehículo':<12} {'Placa':<10} {'Servicio':<20} {'Costo':<12} {'Lavador':<15} {'Estado':<10}
{'-'*80}
"""
            
            if records:
                for record in records:
                    hora = str(record['hora'])[:5] if record['hora'] else '00:00'
                    vehiculo = (record['vehiculo_nombre'] or '')[:11]
                    placa = (record['placa'] or 'N/A')[:9]
                    servicio = (record['servicio_nombre'] or '')[:19]
                    costo = f"${record['costo']:,.0f}" if record['costo'] else '$0'
                    lavador = (record['lavador'] or 'N/A')[:14]
                    estado = (record['pago'] or 'Pendiente')[:9]
                    
                    report_content += f"{hora:<8} {vehiculo:<12} {placa:<10} {servicio:<20} {costo:<12} {lavador:<15} {estado:<10}\n"
            
            text_area.insert('1.0', report_content)
            text_area.configure(state='disabled')
            
            text_area.pack(side='left', fill='both', expand=True)
            scrollbar_v.pack(side='right', fill='y')
            
            buttons_frame = tk.Frame(preview_window, bg='white')
            buttons_frame.pack(fill='x', padx=20, pady=(0, 20))
            
            close_btn = self.create_button(
                buttons_frame, "Cerrar",
                preview_window.destroy, bg_color='#6b7280'
            )
            close_btn.pack(side='left')
            
            def save_report():
                filename = filedialog.asksaveasfilename(
                    title="Guardar Reporte",
                    defaultextension=".txt",
                    filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
                    initialfile=f"reporte_diario_{today}.txt"
                )
                
                if filename:
                    try:
                        with open(filename, 'w', encoding='utf-8') as f:
                            f.write(report_content)
                        messagebox.showinfo("Éxito", 
                            f"Reporte guardado en:\n{filename}")
                    except Exception as e:
                        messagebox.showerror("Error", 
                            f"Error al guardar: {str(e)}")
            
            save_btn = self.create_button(
                buttons_frame, "💾 Guardar",
                save_report, bg_color='#059669'
            )
            save_btn.pack(side='right')
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar reporte: {str(e)}")
    
    def refresh(self):
        """Refrescar datos del cierre de caja"""
        self.load_data()

    def show_context_menu(self, event):
        """Mostrar menú contextual"""
        item = self.records_tree.identify_row(event.y)
        if item:
            self.records_tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)

    def edit_record(self):
        """Editar registro seleccionado - TODOS LOS CAMPOS"""
        selected = self.records_tree.selection()
        if not selected:
            messagebox.showwarning("Advertencia", "Seleccione un registro para editar")
            return
        
        values = self.records_tree.item(selected[0], 'values')
        record_id = values[0]
        
        if not record_id or record_id == '':
            return
        
        # Obtener datos completos del registro
        try:
            query = """
                SELECT r.*, s.nombre as servicio_nombre
                FROM registros r
                LEFT JOIN servicios s ON r.id_servicio = s.id
                WHERE r.id = %s
            """
            record_data = db.execute_query(query, (record_id,))
            if not record_data:
                messagebox.showerror("Error", "No se encontró el registro")
                return
            
            record = record_data[0]
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar registro: {str(e)}")
            return
        
        # Crear ventana de edición
        edit_window = tk.Toplevel(self.parent_frame)
        edit_window.title("Editar Registro Completo")
        edit_window.geometry("500x700")
        edit_window.configure(bg='white')
        edit_window.grab_set()
        
        # Título
        tk.Label(
            edit_window, text=f"Editar Registro #{record_id}",
            font=('Segoe UI', 16, 'bold'), bg='white', fg='#1e293b'
        ).pack(pady=(20, 10))
        
        # Canvas con scroll
        canvas = tk.Canvas(edit_window, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(edit_window, orient="vertical", command=canvas.yview)
        fields_container = tk.Frame(canvas, bg='white')
        
        fields_container.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=fields_container, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=30)
        scrollbar.pack(side="right", fill="y")
        
        # Scroll con mouse (CORREGIDO - sin bind_all)
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind("<MouseWheel>", _on_mousewheel)
        canvas.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
        canvas.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))
        
        # Campos editables
        fields_frame = tk.Frame(fields_container, bg='white')
        fields_frame.pack(fill='both', expand=True, pady=10)
        
        # 1. Fecha
        tk.Label(fields_frame, text="Fecha:", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        fecha_entry = tk.Entry(fields_frame, font=('Segoe UI', 10), relief='solid', borderwidth=1)
        fecha_entry.pack(fill='x', pady=(0, 15), ipady=6)
        fecha_entry.insert(0, str(record['fecha']))
        
        # 2. Hora
        tk.Label(fields_frame, text="Hora:", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        hora_entry = tk.Entry(fields_frame, font=('Segoe UI', 10), relief='solid', borderwidth=1)
        hora_entry.pack(fill='x', pady=(0, 15), ipady=6)
        hora_entry.insert(0, str(record['hora'])[:5] if record['hora'] else '')
        
        # 3. Tipo de Vehículo
        tk.Label(fields_frame, text="Tipo de Vehículo:", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        vehicle_var = tk.StringVar()
        vehicle_options = {
            'motorcycle': 'Motocicleta',
            'car': 'Automóvil',
            'pickup': 'Camioneta',
            'suv': 'SUV',
            'truck': 'Camión'
        }
        vehicle_combo = ttk.Combobox(fields_frame, textvariable=vehicle_var, 
                                    values=list(vehicle_options.values()), 
                                    font=('Segoe UI', 10), state='readonly')
        vehicle_combo.pack(fill='x', pady=(0, 15), ipady=6)
        vehicle_combo.set(vehicle_options.get(record['vehiculo'], 'Automóvil'))
        
        # 4. Placa
        tk.Label(fields_frame, text="Placa:", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        placa_entry = tk.Entry(fields_frame, font=('Segoe UI', 10), relief='solid', borderwidth=1)
        placa_entry.pack(fill='x', pady=(0, 15), ipady=6)
        placa_entry.insert(0, record['placa'] or '')
        
        # 5. Servicio (solo mostrar, no editable)
        tk.Label(fields_frame, text="Servicio:", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        servicio_label = tk.Label(fields_frame, text=record['servicio_nombre'] or 'N/A',
                                font=('Segoe UI', 10), bg='#f9fafb', relief='solid',
                                borderwidth=1, anchor='w', padx=10, pady=8)
        servicio_label.pack(fill='x', pady=(0, 15))
        
        # 6. Costo
        tk.Label(fields_frame, text="Costo ($):", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        costo_entry = tk.Entry(fields_frame, font=('Segoe UI', 10), relief='solid', borderwidth=1)
        costo_entry.pack(fill='x', pady=(0, 15), ipady=6)
        costo_entry.insert(0, str(record['costo']))
        
        # 7. Porcentaje
        tk.Label(fields_frame, text="Porcentaje (%):", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        porcentaje_entry = tk.Entry(fields_frame, font=('Segoe UI', 10), relief='solid', borderwidth=1)
        porcentaje_entry.pack(fill='x', pady=(0, 15), ipady=6)
        porcentaje_entry.insert(0, str(record['porcentaje']))
        
        # 8. Lavador
        tk.Label(fields_frame, text="Lavador:", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        lavador_entry = tk.Entry(fields_frame, font=('Segoe UI', 10), relief='solid', borderwidth=1)
        lavador_entry.pack(fill='x', pady=(0, 15), ipady=6)
        lavador_entry.insert(0, record['lavador'] or '')
        
        # 9. Observaciones
        tk.Label(fields_frame, text="Observaciones:", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        observaciones_text = tk.Text(fields_frame, font=('Segoe UI', 10), relief='solid', 
                                    borderwidth=1, height=3)
        observaciones_text.pack(fill='x', pady=(0, 15))
        observaciones_text.insert('1.0', record['observaciones'] or '')
        
        # 10. Estado de Pago
        tk.Label(fields_frame, text="Estado del Pago:", font=('Segoe UI', 10, 'bold'), bg='white').pack(anchor='w', pady=(0, 5))
        status_var = tk.BooleanVar(value=(record['pago'] == 'Pagado'))
        tk.Checkbutton(fields_frame, text="Pagado", variable=status_var, 
                    font=('Segoe UI', 10), bg='white').pack(anchor='w', pady=(0, 20))
        
        # AHORA SÍ: Definir función on_close DESPUÉS de crear todo
        def on_close():
            canvas.unbind("<MouseWheel>")
            canvas.unbind("<Button-4>")
            canvas.unbind("<Button-5>")
            edit_window.destroy()
        
        # Protocolo de cierre de ventana
        edit_window.protocol("WM_DELETE_WINDOW", on_close)
        
        # Botones
        buttons_frame = tk.Frame(edit_window, bg='white')
        buttons_frame.pack(fill='x', padx=30, pady=20, side='bottom')
        
        tk.Button(
            buttons_frame, text="Cancelar", font=('Segoe UI', 10),
            bg='#6b7280', fg='white', relief='flat', padx=20, pady=10,
            command=on_close
        ).pack(side='left')
        
        def save_all_changes():
            try:
                nueva_fecha = fecha_entry.get().strip()
                nueva_hora = hora_entry.get().strip()
                vehicle_display = vehicle_var.get()
                nuevo_vehiculo = next((k for k, v in vehicle_options.items() if v == vehicle_display), 'car')
                nueva_placa = placa_entry.get().strip().upper() or None
                nuevo_costo = float(costo_entry.get().replace(',', ''))
                nuevo_porcentaje = float(porcentaje_entry.get())
                nuevo_lavador = lavador_entry.get().strip() or None
                nuevas_observaciones = observaciones_text.get('1.0', tk.END).strip() or None
                nuevo_estado = 'Pagado' if status_var.get() else 'Pendiente'
                
                datetime.strptime(nueva_fecha, '%Y-%m-%d')
                datetime.strptime(nueva_hora, '%H:%M')
                
                update_query = """
                    UPDATE registros 
                    SET fecha = %s, hora = %s, vehiculo = %s, placa = %s,
                        costo = %s, porcentaje = %s, lavador = %s,
                        observaciones = %s, pago = %s
                    WHERE id = %s
                """
                
                params = (nueva_fecha, nueva_hora, nuevo_vehiculo, nueva_placa,
                        nuevo_costo, nuevo_porcentaje, nuevo_lavador,
                        nuevas_observaciones, nuevo_estado, record_id)
                
                if db.execute_update(update_query, params):
                    messagebox.showinfo("Éxito", "Registro actualizado correctamente")
                    on_close()
                    self.load_data()
                else:
                    messagebox.showerror("Error", "No se pudo actualizar el registro")
                    
            except ValueError:
                messagebox.showerror("Error", "Formato de fecha/hora o valores numéricos inválidos")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar: {str(e)}")
        
        tk.Button(
            buttons_frame, text="💾 Guardar Cambios", font=('Segoe UI', 10, 'bold'),
            bg='#059669', fg='white', relief='flat', padx=20, pady=10,
            command=save_all_changes
        ).pack(side='right')
    def delete_record(self):
        """Eliminar registro seleccionado"""
        selected = self.records_tree.selection()
        if not selected:
            messagebox.showwarning("Advertencia", "Seleccione un registro")
            return
        
        values = self.records_tree.item(selected[0], 'values')
        record_id = values[0]
        
        if not record_id:
            return
        
        if messagebox.askyesno("Confirmar", f"¿Eliminar el registro #{record_id}?"):
            try:
                if db.execute_update("DELETE FROM registros WHERE id = %s", (record_id,)):
                    messagebox.showinfo("Éxito", "Registro eliminado")
                    self.load_data()
                else:
                    messagebox.showerror("Error", "No se pudo eliminar")
            except Exception as e:
                messagebox.showerror("Error", f"Error: {str(e)}")