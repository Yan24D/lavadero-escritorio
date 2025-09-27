"""
Módulo de Historial Completo - Visualización y Análisis de Registros
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import calendar
from .base_module import BaseModule

# Para exportar a Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class HistorialModule(BaseModule):
    """Módulo completo de historial con filtros avanzados y exportación"""
    
    def setup_module(self):
        """Configurar módulo de historial"""
        # Estado inicial de filtros
        self.current_filters = {
            'fecha_inicio': None,
            'fecha_fin': None,
            'lavador': 'Todos',
            'vehiculo': 'Todos',
            'servicio': 'Todos',
            'estado_pago': 'Todos'
        }
        
        # Configurar interfaz
        self.create_header()
        self.create_filters_section()
        self.create_results_section()
        self.create_statistics_section()
        
        # Cargar datos iniciales
        self.load_filter_options()
        self.load_historial_data()
    
    def create_header(self):
        """Crear header del módulo"""
        header_frame = tk.Frame(self.parent, bg='#f8fafc')
        header_frame.pack(fill='x', pady=(0, 20))
        
        title_label = tk.Label(
            header_frame,
            text="📋 Historial Completo de Servicios",
            font=('Segoe UI', 24, 'bold'),
            fg='#1e293b',
            bg='#f8fafc'
        )
        title_label.pack(side='left')
        
        # Botones de acción
        actions_frame = tk.Frame(header_frame, bg='#f8fafc')
        actions_frame.pack(side='right')
        
        # Botón Exportar Excel
        export_btn = tk.Button(
            actions_frame,
            text="📊 Exportar Excel" if EXCEL_AVAILABLE else "📊 Excel N/D",
            font=('Segoe UI', 12, 'bold'),
            bg='#059669' if EXCEL_AVAILABLE else '#9ca3af',
            fg='white',
            relief='flat',
            padx=20,
            pady=10,
            cursor='hand2' if EXCEL_AVAILABLE else 'not-allowed',
            state='normal' if EXCEL_AVAILABLE else 'disabled',
            command=self.export_to_excel
        )
        export_btn.pack(side='right', padx=(10, 0))
        
        # Botón Limpiar Filtros
        clear_btn = tk.Button(
            actions_frame,
            text="🔄 Limpiar Filtros",
            font=('Segoe UI', 12, 'bold'),
            bg='#6b7280',
            fg='white',
            relief='flat',
            padx=20,
            pady=10,
            cursor='hand2',
            command=self.clear_filters
        )
        clear_btn.pack(side='right')
    
    def create_filters_section(self):
        """Crear sección de filtros avanzados"""
        filters_frame = tk.Frame(self.parent, bg='white', relief='solid', borderwidth=1)
        filters_frame.pack(fill='x', pady=(0, 20))
        
        # Header de filtros
        filters_header = tk.Frame(filters_frame, bg='#f8fafc')
        filters_header.pack(fill='x', padx=1, pady=1)
        
        tk.Label(
            filters_header,
            text="🔍 Filtros de Búsqueda Avanzada",
            font=('Segoe UI', 16, 'bold'),
            fg='#1e293b',
            bg='#f8fafc'
        ).pack(side='left', padx=20, pady=15)
        
        # Container de filtros
        filters_container = tk.Frame(filters_frame, bg='white')
        filters_container.pack(fill='x', padx=20, pady=20)
        
        # Primera fila de filtros
        row1 = tk.Frame(filters_container, bg='white')
        row1.pack(fill='x', pady=(0, 15))
        
        # Filtro de fechas
        date_frame = tk.Frame(row1, bg='white')
        date_frame.pack(side='left', padx=(0, 30))
        
        tk.Label(
            date_frame,
            text="Rango de Fechas:",
            font=('Segoe UI', 11, 'bold'),
            bg='white',
            fg='#374151'
        ).pack(anchor='w')
        
        date_inputs = tk.Frame(date_frame, bg='white')
        date_inputs.pack(fill='x', pady=(5, 0))
        
        # Fecha inicio
        tk.Label(date_inputs, text="Desde:", font=('Segoe UI', 10), bg='white').pack(side='left')
        self.fecha_inicio_var = tk.StringVar()
        fecha_inicio_entry = tk.Entry(
            date_inputs,
            textvariable=self.fecha_inicio_var,
            font=('Segoe UI', 10),
            width=12,
            relief='solid',
            borderwidth=1
        )
        fecha_inicio_entry.pack(side='left', padx=(5, 10))
        
        # Fecha fin
        tk.Label(date_inputs, text="Hasta:", font=('Segoe UI', 10), bg='white').pack(side='left')
        self.fecha_fin_var = tk.StringVar()
        fecha_fin_entry = tk.Entry(
            date_inputs,
            textvariable=self.fecha_fin_var,
            font=('Segoe UI', 10),
            width=12,
            relief='solid',
            borderwidth=1
        )
        fecha_fin_entry.pack(side='left', padx=5)
        
        # Botón fechas rápidas
        quick_dates_btn = tk.Button(
            date_inputs,
            text="📅",
            font=('Segoe UI', 10),
            bg='#2563eb',
            fg='white',
            relief='flat',
            cursor='hand2',
            command=self.show_quick_dates
        )
        quick_dates_btn.pack(side='left', padx=(10, 0))
        
        # Filtro de Lavador
        lavador_frame = tk.Frame(row1, bg='white')
        lavador_frame.pack(side='left', padx=(0, 30))
        
        tk.Label(
            lavador_frame,
            text="Lavador:",
            font=('Segoe UI', 11, 'bold'),
            bg='white',
            fg='#374151'
        ).pack(anchor='w')
        
        self.lavador_var = tk.StringVar(value="Todos")
        self.lavador_combo = ttk.Combobox(
            lavador_frame,
            textvariable=self.lavador_var,
            state='readonly',
            width=18
        )
        self.lavador_combo.pack(pady=(5, 0))
        
        # Segunda fila de filtros
        row2 = tk.Frame(filters_container, bg='white')
        row2.pack(fill='x', pady=(0, 15))
        
        # Filtro de Tipo de Vehículo
        vehiculo_frame = tk.Frame(row2, bg='white')
        vehiculo_frame.pack(side='left', padx=(0, 30))
        
        tk.Label(
            vehiculo_frame,
            text="Tipo de Vehículo:",
            font=('Segoe UI', 11, 'bold'),
            bg='white',
            fg='#374151'
        ).pack(anchor='w')
        
        self.vehiculo_var = tk.StringVar(value="Todos")
        self.vehiculo_combo = ttk.Combobox(
            vehiculo_frame,
            textvariable=self.vehiculo_var,
            values=['Todos', 'Motocicleta', 'Automóvil', 'Camioneta', 'SUV', 'Camión'],
            state='readonly',
            width=18
        )
        self.vehiculo_combo.pack(pady=(5, 0))
        
        # Filtro de Servicio
        servicio_frame = tk.Frame(row2, bg='white')
        servicio_frame.pack(side='left', padx=(0, 30))
        
        tk.Label(
            servicio_frame,
            text="Servicio:",
            font=('Segoe UI', 11, 'bold'),
            bg='white',
            fg='#374151'
        ).pack(anchor='w')
        
        self.servicio_var = tk.StringVar(value="Todos")
        self.servicio_combo = ttk.Combobox(
            servicio_frame,
            textvariable=self.servicio_var,
            state='readonly',
            width=18
        )
        self.servicio_combo.pack(pady=(5, 0))
        
        # Filtro de Estado de Pago
        estado_frame = tk.Frame(row2, bg='white')
        estado_frame.pack(side='left')
        
        tk.Label(
            estado_frame,
            text="Estado de Pago:",
            font=('Segoe UI', 11, 'bold'),
            bg='white',
            fg='#374151'
        ).pack(anchor='w')
        
        self.estado_var = tk.StringVar(value="Todos")
        self.estado_combo = ttk.Combobox(
            estado_frame,
            textvariable=self.estado_var,
            values=['Todos', 'Pagado', 'Pendiente'],
            state='readonly',
            width=15
        )
        self.estado_combo.pack(pady=(5, 0))
        
        # Botón Aplicar Filtros
        apply_frame = tk.Frame(filters_container, bg='white')
        apply_frame.pack(pady=(15, 0))
        
        apply_btn = tk.Button(
            apply_frame,
            text="🔍 APLICAR FILTROS",
            font=('Segoe UI', 12, 'bold'),
            bg='#2563eb',
            fg='white',
            relief='flat',
            padx=30,
            pady=12,
            cursor='hand2',
            command=self.apply_filters
        )
        apply_btn.pack()
        
        # Bind eventos para aplicar filtros automáticamente
        for combo in [self.lavador_combo, self.vehiculo_combo, self.servicio_combo, self.estado_combo]:
            combo.bind('<<ComboboxSelected>>', lambda e: self.apply_filters())
        
        for entry in [fecha_inicio_entry, fecha_fin_entry]:
            entry.bind('<KeyRelease>', lambda e: self.parent.after(1000, self.apply_filters))
    
    def create_results_section(self):
        """Crear sección de resultados"""
        results_frame = tk.Frame(self.parent, bg='white', relief='solid', borderwidth=1)
        results_frame.pack(fill='both', expand=True, pady=(0, 20))
        
        # Header de resultados
        results_header = tk.Frame(results_frame, bg='#f8fafc')
        results_header.pack(fill='x', padx=1, pady=1)
        
        self.results_title = tk.Label(
            results_header,
            text="📊 Resultados del Historial (Cargando...)",
            font=('Segoe UI', 16, 'bold'),
            fg='#1e293b',
            bg='#f8fafc'
        )
        self.results_title.pack(side='left', padx=20, pady=15)
        
        # Tabla de resultados
        table_container = tk.Frame(results_frame, bg='white')
        table_container.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Configurar Treeview
        columns = ('ID', 'Fecha', 'Hora', 'Vehículo', 'Placa', 'Servicio', 'Lavador', 'Costo', 'Comisión', 'Estado')
        
        self.results_tree = ttk.Treeview(table_container, columns=columns, show='headings', height=15)
        
        # Configurar columnas
        column_configs = [
            ('ID', 50, 'center'),
            ('Fecha', 90, 'center'),
            ('Hora', 70, 'center'),
            ('Vehículo', 100, 'center'),
            ('Placa', 80, 'center'),
            ('Servicio', 200, 'w'),
            ('Lavador', 150, 'w'),
            ('Costo', 100, 'center'),
            ('Comisión', 100, 'center'),
            ('Estado', 80, 'center')
        ]
        
        for col, width, anchor in column_configs:
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=width, anchor=anchor)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(table_container, orient='vertical', command=self.results_tree.yview)
        h_scrollbar = ttk.Scrollbar(table_container, orient='horizontal', command=self.results_tree.xview)
        
        self.results_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack elementos
        self.results_tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        table_container.grid_rowconfigure(0, weight=1)
        table_container.grid_columnconfigure(0, weight=1)
        
        # Tags para colores
        self.results_tree.tag_configure('pagado', background='#f0fdf4')
        self.results_tree.tag_configure('pendiente', background='#fef2f2')
    
    def create_statistics_section(self):
        """Crear sección de estadísticas de filtros"""
        stats_frame = tk.Frame(self.parent, bg='white', relief='solid', borderwidth=1)
        stats_frame.pack(fill='x')
        
        # Header de estadísticas
        stats_header = tk.Frame(stats_frame, bg='#f8fafc')
        stats_header.pack(fill='x', padx=1, pady=1)
        
        tk.Label(
            stats_header,
            text="📈 Estadísticas de los Resultados Filtrados",
            font=('Segoe UI', 16, 'bold'),
            fg='#1e293b',
            bg='#f8fafc'
        ).pack(side='left', padx=20, pady=15)
        
        # Container de estadísticas
        stats_container = tk.Frame(stats_frame, bg='white')
        stats_container.pack(fill='x', padx=20, pady=20)
        
        # Grid de estadísticas
        self.stats_labels = {}
        
        stats_configs = [
            ('total_registros', 'Total Registros', '#2563eb'),
            ('total_ingresos', 'Ingresos Totales', '#059669'),
            ('total_comisiones', 'Comisiones Totales', '#d97706'),
            ('ganancia_neta', 'Ganancia Neta', '#059669'),
            ('promedio_servicio', 'Promedio por Servicio', '#7c3aed'),
            ('servicios_pendientes', 'Servicios Pendientes', '#dc2626')
        ]
        
        for i, (key, title, color) in enumerate(stats_configs):
            # Calcular posición en grid
            row = i // 3
            col = i % 3
            
            stat_frame = tk.Frame(stats_container, bg=color, relief='solid', borderwidth=1)
            stat_frame.grid(row=row, column=col, padx=10, pady=10, sticky='ew')
            
            # Configurar columnas del grid
            if col == 0:
                stats_container.grid_columnconfigure(0, weight=1)
                stats_container.grid_columnconfigure(1, weight=1)
                stats_container.grid_columnconfigure(2, weight=1)
            
            # Contenido del stat
            stat_content = tk.Frame(stat_frame, bg=color)
            stat_content.pack(fill='both', expand=True, padx=15, pady=12)
            
            # Título
            tk.Label(
                stat_content,
                text=title,
                font=('Segoe UI', 10, 'bold'),
                fg='white',
                bg=color
            ).pack()
            
            # Valor
            value_label = tk.Label(
                stat_content,
                text="0",
                font=('Segoe UI', 14, 'bold'),
                fg='white',
                bg=color
            )
            value_label.pack(pady=(5, 0))
            
            self.stats_labels[key] = value_label
    
    def show_quick_dates(self):
        """Mostrar opciones de fechas rápidas"""
        quick_menu = tk.Menu(self.parent, tearoff=0)
        
        today = datetime.now()
        
        # Opciones de fechas rápidas
        date_options = [
            ("Hoy", today, today),
            ("Ayer", today - timedelta(days=1), today - timedelta(days=1)),
            ("Últimos 7 días", today - timedelta(days=7), today),
            ("Últimos 30 días", today - timedelta(days=30), today),
            ("Este mes", today.replace(day=1), today),
            ("Mes anterior", (today.replace(day=1) - timedelta(days=1)).replace(day=1), today.replace(day=1) - timedelta(days=1)),
            ("Este año", today.replace(month=1, day=1), today)
        ]
        
        for label, fecha_inicio, fecha_fin in date_options:
            quick_menu.add_command(
                label=label,
                command=lambda fi=fecha_inicio, ff=fecha_fin: self.set_quick_dates(fi, ff)
            )
        
        # Mostrar menú en la posición del cursor
        try:
            quick_menu.post(self.parent.winfo_pointerx(), self.parent.winfo_pointery())
        except:
            pass
    
    def set_quick_dates(self, fecha_inicio, fecha_fin):
        """Establecer fechas rápidas"""
        self.fecha_inicio_var.set(fecha_inicio.strftime('%Y-%m-%d'))
        self.fecha_fin_var.set(fecha_fin.strftime('%Y-%m-%d'))
        self.apply_filters()
    
    def load_filter_options(self):
        """Cargar opciones para los filtros"""
        try:
            # Cargar lavadores
            lavadores_query = "SELECT DISTINCT lavador FROM vista_registros_completos ORDER BY lavador"
            lavadores_result = self.db.execute_query(lavadores_query)
            
            lavadores = ['Todos']
            if lavadores_result:
                lavadores.extend([row['lavador'] for row in lavadores_result])
            
            self.lavador_combo['values'] = lavadores
            
            # Cargar servicios
            servicios_query = "SELECT DISTINCT servicio_nombre FROM vista_registros_completos ORDER BY servicio_nombre"
            servicios_result = self.db.execute_query(servicios_query)
            
            servicios = ['Todos']
            if servicios_result:
                servicios.extend([row['servicio_nombre'] for row in servicios_result])
            
            self.servicio_combo['values'] = servicios
            
        except Exception as e:
            print(f"Error cargando opciones de filtros: {e}")
    
    def clear_filters(self):
        """Limpiar todos los filtros"""
        self.fecha_inicio_var.set("")
        self.fecha_fin_var.set("")
        self.lavador_var.set("Todos")
        self.vehiculo_var.set("Todos")
        self.servicio_var.set("Todos")
        self.estado_var.set("Todos")
        
        self.current_filters = {
            'fecha_inicio': None,
            'fecha_fin': None,
            'lavador': 'Todos',
            'vehiculo': 'Todos',
            'servicio': 'Todos',
            'estado_pago': 'Todos'
        }
        
        self.load_historial_data()
    
    def apply_filters(self):
        """Aplicar filtros seleccionados"""
        # Actualizar filtros actuales
        self.current_filters['fecha_inicio'] = self.fecha_inicio_var.get().strip() or None
        self.current_filters['fecha_fin'] = self.fecha_fin_var.get().strip() or None
        self.current_filters['lavador'] = self.lavador_var.get()
        self.current_filters['vehiculo'] = self.vehiculo_var.get()
        self.current_filters['servicio'] = self.servicio_var.get()
        self.current_filters['estado_pago'] = self.estado_var.get()
        
        # Cargar datos filtrados
        self.load_historial_data()
    
    def load_historial_data(self):
        """Cargar datos del historial con filtros aplicados"""
        try:
            # Limpiar tabla
            for item in self.results_tree.get_children():
                self.results_tree.delete(item)
            
            # Construir query base
            base_query = """
                SELECT 
                    id, fecha, hora, vehiculo_nombre, placa, servicio_nombre,
                    lavador, costo, comision_calculada, pago
                FROM vista_registros_completos
                WHERE 1=1
            """
            
            params = []
            
            # Aplicar filtros
            if self.current_filters['fecha_inicio']:
                base_query += " AND fecha >= %s"
                params.append(self.current_filters['fecha_inicio'])
            
            if self.current_filters['fecha_fin']:
                base_query += " AND fecha <= %s"
                params.append(self.current_filters['fecha_fin'])
            
            if self.current_filters['lavador'] != 'Todos':
                base_query += " AND lavador = %s"
                params.append(self.current_filters['lavador'])
            
            if self.current_filters['vehiculo'] != 'Todos':
                # Mapear nombres a valores de BD
                vehiculo_map = {
                    'Motocicleta': 'motorcycle',
                    'Automóvil': 'car',
                    'Camioneta': 'pickup',
                    'SUV': 'suv',
                    'Camión': 'truck'
                }
                if self.current_filters['vehiculo'] in vehiculo_map:
                    base_query += " AND vehiculo = %s"
                    params.append(vehiculo_map[self.current_filters['vehiculo']])
            
            if self.current_filters['servicio'] != 'Todos':
                base_query += " AND servicio_nombre = %s"
                params.append(self.current_filters['servicio'])
            
            if self.current_filters['estado_pago'] != 'Todos':
                base_query += " AND pago = %s"
                params.append(self.current_filters['estado_pago'])
            
            base_query += " ORDER BY fecha DESC, hora DESC"
            
            # Ejecutar query
            results = self.db.execute_query(base_query, params if params else None)
            
            if results:
                for row in results:
                    # Determinar tag por estado
                    tag = 'pagado' if row['pago'] == 'Pagado' else 'pendiente'
                    
                    self.results_tree.insert('', 'end', values=(
                        row['id'],
                        row['fecha'].strftime('%d/%m/%Y') if hasattr(row['fecha'], 'strftime') else str(row['fecha']),
                        str(row['hora']),
                        row['vehiculo_nombre'],
                        row['placa'],
                        row['servicio_nombre'],
                        row['lavador'],
                        f"${row['costo']:,.0f}",
                        f"${row['comision_calculada']:,.0f}",
                        row['pago']
                    ), tags=(tag,))
            
            # Actualizar título con contador
            total_records = len(results) if results else 0
            self.results_title.config(text=f"📊 Resultados del Historial ({total_records:,} registros)")
            
            # Actualizar estadísticas
            self.update_statistics(results)
            
        except Exception as e:
            print(f"Error cargando historial: {e}")
            self.results_title.config(text="📊 Error al cargar historial")
            messagebox.showerror("Error", "Error al cargar el historial de servicios")
    
    def update_statistics(self, data):
        """Actualizar estadísticas basadas en los datos filtrados"""
        if not data:
            # Resetear estadísticas si no hay datos
            for key in self.stats_labels:
                self.stats_labels[key].config(text="0")
            return
        
        # Calcular estadísticas
        total_registros = len(data)
        total_ingresos = sum(row['costo'] for row in data)
        total_comisiones = sum(row['comision_calculada'] for row in data)
        ganancia_neta = total_ingresos - total_comisiones
        promedio_servicio = total_ingresos / total_registros if total_registros > 0 else 0
        servicios_pendientes = sum(1 for row in data if row['pago'] == 'Pendiente')
        
        # Actualizar labels
        self.stats_labels['total_registros'].config(text=f"{total_registros:,}")
        self.stats_labels['total_ingresos'].config(text=f"${total_ingresos:,.0f}")
        self.stats_labels['total_comisiones'].config(text=f"${total_comisiones:,.0f}")
        self.stats_labels['ganancia_neta'].config(text=f"${ganancia_neta:,.0f}")
        self.stats_labels['promedio_servicio'].config(text=f"${promedio_servicio:,.0f}")
        self.stats_labels['servicios_pendientes'].config(text=f"{servicios_pendientes}")
    
    def export_to_excel(self):
        """Exportar datos filtrados a Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", "La librería openpyxl no está instalada.\nInstale con: pip install openpyxl")
            return
        
        try:
            # Solicitar ubicación del archivo
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Guardar historial como Excel"
            )
            
            if not filename:
                return
            
            # Obtener datos actuales de la tabla
            items = self.results_tree.get_children()
            if not items:
                messagebox.showwarning("Advertencia", "No hay datos para exportar")
                return
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historial de Servicios"
            
            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Headers
            headers = ['ID', 'Fecha', 'Hora', 'Tipo Vehículo', 'Placa', 'Servicio', 'Lavador', 'Costo', 'Comisión', 'Estado']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Datos
            for row_idx, item in enumerate(items, 2):
                values = self.results_tree.item(item, 'values')
                for col_idx, value in enumerate(values, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Ajustar ancho de columnas
            column_widths = [5, 12, 10, 15, 10, 30, 25, 12, 12, 10]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # Agregar hoja de estadísticas
            stats_ws = wb.create_sheet(title="Estadísticas")
            
            # Títulos de estadísticas
            stats_data = [
                ("Estadísticas del Historial Filtrado", ""),
                ("", ""),
                ("Total de Registros", f"{len(items):,}"),
                ("Total Ingresos", self.stats_labels['total_ingresos'].cget('text')),
                ("Total Comisiones", self.stats_labels['total_comisiones'].cget('text')),
                ("Ganancia Neta", self.stats_labels['ganancia_neta'].cget('text')),
                ("Promedio por Servicio", self.stats_labels['promedio_servicio'].cget('text')),
                ("Servicios Pendientes", self.stats_labels['servicios_pendientes'].cget('text')),
                ("", ""),
                ("Filtros Aplicados:", ""),
                ("Fecha Inicio", self.current_filters['fecha_inicio'] or "No aplicado"),
                ("Fecha Fin", self.current_filters['fecha_fin'] or "No aplicado"),
                ("Lavador", self.current_filters['lavador']),
                ("Tipo de Vehículo", self.current_filters['vehiculo']),
                ("Servicio", self.current_filters['servicio']),
                ("Estado de Pago", self.current_filters['estado_pago']),
                ("", ""),
                ("Fecha de Exportación", datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
            ]
            
            for row_idx, (label, value) in enumerate(stats_data, 1):
                stats_ws.cell(row=row_idx, column=1, value=label)
                stats_ws.cell(row=row_idx, column=2, value=value)
                
                # Estilo para títulos
                if label == "Estadísticas del Historial Filtrado":
                    stats_ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
                elif label and not value:
                    stats_ws.cell(row=row_idx, column=1).font = Font(bold=True)
            
            # Ajustar columnas de estadísticas
            stats_ws.column_dimensions['A'].width = 25
            stats_ws.column_dimensions['B'].width = 20
            
            # Guardar archivo
            wb.save(filename)
            
            messagebox.showinfo(
                "Exportación Exitosa",
                f"Historial exportado exitosamente a:\n{filename}\n\n"
                f"Se exportaron {len(items):,} registros con estadísticas incluidas."
            )
            
        except Exception as e:
            print(f"Error exportando a Excel: {e}")
            messagebox.showerror("Error", f"Error al exportar a Excel:\n{str(e)}")
    
    def refresh(self):
        """Refrescar datos del módulo"""
        self.load_filter_options()
        self.load_historial_data()
    
    def cleanup(self):
        """Limpiar recursos del módulo"""
        pass